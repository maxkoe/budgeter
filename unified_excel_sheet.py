import json
import pandas as pd
import numpy as np
import openpyxl as pyxl
import re
from sqlite3 import IntegrityError

from excel_helpers import *
from db_entries import *

from itertools import chain
from warnings import warn as Warning

###############################################################################
# creating a unified excel sheet based on definition data from a .json-file
###############################################################################

def create_excel_table_from_data(dataframe, filename, db, definition_data) :
    
    if type(definition_data) is str : 
        with open(definition_data) as f :
            json_data = f.read()
            sheet_def = json.loads(json_data)
    elif type(definition_data) is dict :
        sheet_def = definition_data
    else :
        raise ValueError(
            "Unsupported data type given for argument definition_data")
        
    main_sheet_def_dict = sheet_def['main_sheet']
    lookup_sheets_defs = sheet_def['lookup_sheets']
    type_style_defs = sheet_def['type_styles']
    
    main_sheet_def = pd.DataFrame.from_dict(main_sheet_def_dict, 
                                            orient='index').sort_values(
                                                    by='excel_col')
    
    ## Insert the missing cols into the dataframe with the defined default 
    ## value
    for col_title, col_props in main_sheet_def.iterrows() :
        if col_title not in dataframe.columns :
            dataframe[col_title] = col_props['default']
            
    dataframe = dataframe[main_sheet_def.index.tolist()]

    ## Save the dataframe to excel and open it with openpyxl
    dataframe.to_excel(filename)
    wb = pyxl.load_workbook(filename, data_only=False)
    the_sheet = wb.active
    
    ## Add data validation to the excel document
    #### First enforce Checkmarks in all cols in the list checkmark_cols
    
    validate_checkmark = pyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1='"X"', allow_blank=True)
    validate_checkmark.error ='Your entry is not in the list'
    validate_checkmark.errorTitle = 'Invalid Entry'
    validate_checkmark.prompt = 'Please check with an X or leave blank'
    validate_checkmark.promptTitle = 'Checkmark'
    
    for col_title, col_props in main_sheet_def[
                            main_sheet_def['type'] == 'checkmark'].iterrows() :
        validate_checkmark.ranges.append(col_props['excel_col'] + '2:' +
                                         col_props['excel_col'] + 
                                         str(the_sheet.max_row))
  
    the_sheet.add_data_validation(validate_checkmark)
    
    ## Add the lookup tables
    def add_lookup_sheet(wb, db, db_table, db_cols, sheet_name = 'lookup', 
                         sorted_by = None) :
        crsr = db.cursor()
        if sorted_by is None : 
            crsr.execute('SELECT {} FROM {}'.format(
                ','.join(db_cols), db_table))
        else :
            crsr.execute('SELECT {} FROM {} ORDER BY {}'.format(
                ','.join(db_cols), db_table, sorted_by))
        results = crsr.fetchall()
        
        lookup_sheet = wb.create_sheet(sheet_name)
        for row in results : 
            lookup_sheet.append(row)
    
    for lookup_sheet_def in lookup_sheets_defs :
        add_lookup_sheet(wb, db, **lookup_sheet_def)
        
    
    ## Insert some lookup tables for data validation and to make the excel 
    ## document more readable
    #### Some functions to make comfortable data validation        
    def add_validation_default_notes(dv) :
        dv.error ='Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select from the list'
        dv.promptTitle = 'List Selection'
        return dv  
    
    ###### Event Types
    for col_title, col_props in main_sheet_def[
            main_sheet_def['type'] == 'validated'].iterrows() :
        validation = pyxl.worksheet.datavalidation.DataValidation(
            **col_props['validation'])
        validation.ranges.append(
            col_props['excel_col'] + '2:' + col_props['excel_col'] + 
            str(the_sheet.max_row))
        the_sheet.add_data_validation(validation)
    
    for i in range(2, the_sheet.max_row + 1) :
        for col_title, col_props in main_sheet_def.iterrows() :
            if col_props['type'] in type_style_defs.keys() :
                styles = type_style_defs[col_props['type']]
                for style, args in styles.items() :
                    if style == 'PatternFill' :
                        the_fill = pyxl.styles.PatternFill(**args)
                        the_sheet[col_props[
                            'excel_col'] + str(i)].fill = the_fill
                    elif style == 'number_format' :
                        the_sheet[col_props[
                            'excel_col'] + str(i)].number_format = args
                    else :
                        raise RuntimeError("Undefined Style used")
            if col_props['type'] == 'calculated' :
                row_ex = re.compile(r"\{\{ROW}}")
                the_formula_with_row = row_ex.sub(str(i), col_props['formula'])

                col_ex = re.compile(r"\{\{.*?}}")
                matches = col_ex.finditer(the_formula_with_row)
                for match in matches : 
                    key_name_ex = re.compile('[a-z_]+')
                    dict_item = sheet_def
                    for key in key_name_ex.finditer(match.group()) :
                        dict_item = dict_item[key.group()]
                the_formula = col_ex.sub(dict_item, the_formula_with_row)
                the_sheet[col_props['excel_col'] + str(i)] = the_formula

    left_align = pyxl.styles.Alignment(horizontal='left')
    for col_title, col_props in main_sheet_def.iterrows() :
        the_sheet[col_props['excel_col'] + '1'].alignment = left_align

    for column_cells in the_sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        the_sheet.column_dimensions[column_cells[0].column].width = length

    for col_title, col_props in main_sheet_def.iterrows() :
        if col_props['type'] == 'checkmark' :
            length = 4
        elif col_props['type'] in ['validated', 'free', 'calculated'] :
            length = 10
        else :
            length = max(len(str(cell.value)) for cell in 
                            the_sheet[col_props['excel_col']])
        the_sheet.column_dimensions[col_props['excel_col']].width = length
    
    for col_title, col_props in main_sheet_def[
            main_sheet_def['hidden'] == True].iterrows() :
        the_sheet.column_dimensions[col_props['excel_col']].hidden = True    
    
    wb.save(filename)

###############################################################################
# some helper functions to easily generate comments for the db and excel files
###############################################################################

def autogenerate_database_comment_multiple_excel_ranges(filename, ranges,
                                                        comment_sheet) :
    try :
        ranges_string = ', '.join(ranges[:-1]) + \
                        (' and ' if len(ranges) > 1 else '') + ranges[-1]
    except IndexError :
        raise ValueError('You should provide at least one Excel range')
    return 'This entry was automatically generated from the excel file ' + \
           '{}. It is based on the cells {}. A note has been '.format(
               filename, ranges_string) + \
           'added to the respective cells in the sheet {}.'.format(comment_sheet.title)

def autogenerate_database_comment(filename, range_, comment_sheet) :
    return autogenerate_database_comment_multiple_excel_ranges(
                filename, [range_], comment_sheet)

def autogenerate_excel_comment(date, db_filename, the_id) :
    return 'On {} this cell was automatically read and '.format(date) + \
           'inserted into the database {}. The id ofthe entry is {}.'.format(
               date, db_filename, the_id)

###############################################################################
# Parse the unified excel sheet
###############################################################################

def parse_unified_excel_table(dataframe, db, db_filename, current_sheet, 
			      comment_sheet, excel_filename, baseframe = None):
    if baseframe is None :
        baseframe = dataframe.copy()
    untreated_data = dataframe[dataframe['treated'] == 'XX']
    
    event_groups_lookup_df = pd.read_sql('SELECT * FROM event_groups', db)
    temporary_id_lookup_dict = {}

    for index, row in dataframe.iterrows() :
        if row['ignore'] == 'X' :
            untreated_data = untreated_data.append(dataframe.loc[index])
            continue
        if row['treated'] == 'Yes' :
            Warning('A treated data row was encountered. This row will ' + 
		    'be ignored.')
            Warning("It is recommended to set the 'ignore'-row to 'X' for " +
		    "treated rows.")
            continue

        excel_ranges = ([row['excel_range_x']] if type(row['excel_range_x'])
                            is str else []) + \
                       ([row['excel_range_y']] if type(row['excel_range_y'])
                            is str else [])

        if not np.isnan(row['temporary_id']) and \
                row['temporary_id'] in temporary_id_lookup_dict :
            the_id = temporary_id_lookup_dict[row['temporary_id']]
        else :
            the_id = create_money_event(db, row['type_description'], 
                                        row['description'], row['date'])


            full_excel_ranges = [current_sheet.title + '!' + a_range 
                                 for a_range in excel_ranges]

            database_comment = \
                autogenerate_database_comment_multiple_excel_ranges(
                    excel_filename, full_excel_ranges, comment_sheet.title)

            add_database_event(db, the_id, database_comment)

            if not np.isnan(row['temporary_id']) :
                temporary_id_lookup_dict[row['temporary_id']] = the_id

        ## ToDo : Add commenting into database and Excel

        if row['is_budget_event'] == 'X' :
            add_budget_event(db, the_id, row['budget_pot'], row['amount'])
        if row['is_payment'] == 'X' :
            add_payment(db, the_id, row['money_pot'], row['amount'])
        if row['is_recieving'] == 'X' :
            add_payment(db, the_id, row['money_pot'], row['amount'])
        if row['is_transfer'] == 'X' :
            add_payment(db, the_id, row['money_pot'], row['money_pot_sink'], 
                        row['amount'])

        if not np.isnan(row['in_group']) :  
            group_id = int(row['in_group'])
            if row['in_group'] not in event_groups_lookup_df['group_id'] :
                if row['group_name'][0] == '=' :
                    raise ValueError("You should not use a formula as group " +
                                     "name! The given group name was {}.".format(
                                        row['group_name']))
                crsr = db.cursor()
                crsr.execute('INSERT INTO event_groups VALUES ({}, "{}")'.format(
                    row['in_group'], row['group_name']))
                db.commit()
                event_groups_lookup_df = pd.read_sql(
                                            'SELECT * FROM event_groups', db)
            try : 
                crsr = db.cursor()
                crsr.execute('INSERT INTO event_in_group VALUES ' + 
                             '({}, {})'.format(row['in_group'], the_id))
                db.commit()
            except IntegrityError as IE :
                if 'UNIQUE constraint failed' in str(IE) :
                    Warning(str(IE))
                else :
                    raise IE
            except :
                raise

        excel_comment = autogenerate_excel_comment(
                dt.now().strftime('%Y-%m-%d'), db_filename, the_id)
        excel_cell_list = list(chain.from_iterable(
                [list_from_range_string(a_range) for a_range in excel_ranges]))
        put_comment_into_excel(comment_sheet, excel_cell_list, excel_comment)

        baseframe.loc[index, 'treated'] = 'Yes'
    
    return untreated_data
